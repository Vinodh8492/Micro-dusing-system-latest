from flask import Flask, Blueprint  , request,current_app, jsonify # type: ignore
from extensions import db
from models.material import Material, MaterialTransaction, MaterialSchema, MaterialTransactionSchema
from sqlalchemy.exc import IntegrityError # type: ignore
from flask import send_file
import io
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
import barcode
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import re
import tempfile


material_bp = Blueprint("materials", __name__)

material_schema = MaterialSchema()
materials_schema = MaterialSchema(many=True)

transaction_schema = MaterialTransactionSchema()
transactions_schema = MaterialTransactionSchema(many=True)

@material_bp.route("/materials/export/barcodes", methods=["GET"])
def export_materials_excel_with_barcodes():
    try:
        materials = Material.query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Material Barcodes"
        ws.append(["Title", "Barcode ID", "Scannable Barcode"])

        row_number = 2

        for material in materials:
            if material.barcode_id:
                barcode_id = material.barcode_id

                try:
                    # Create a temp directory and save barcode
                    temp_dir = tempfile.gettempdir()
                    barcode_filename = f"{barcode_id}"
                    barcode_path = os.path.join(temp_dir, f"{barcode_filename}.png")

                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(os.path.join(temp_dir, barcode_filename))  # Don't include .png here

                    # Resize for Excel
                    image = PILImage.open(barcode_path)
                    image = image.resize((200, 60))
                    image.save(barcode_path)

                    # Fill Excel
                    ws.cell(row=row_number, column=1, value=material.title)
                    ws.cell(row=row_number, column=2, value=barcode_id)

                    img = ExcelImage(barcode_path)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"C{row_number}")

                    # Optional: delete temp image after
                    os.remove(barcode_path)

                    row_number += 1

                except Exception as e:
                    print(f"Failed to generate barcode for {barcode_id}: {e}")
                    # Still add text if image fails
                    ws.cell(row=row_number, column=1, value=material.title)
                    ws.cell(row=row_number, column=2, value=barcode_id)
                    row_number += 1

        # Save to bytes
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="materials_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ➤ Create a new Material
@material_bp.route("/materials", methods=["POST"])
def add_material():
    try:
        data = request.get_json()

        new_material = Material(
            title=data.get("title"),
            description=data.get("description"),
            unit_of_measure=data.get("unit_of_measure"),
            current_quantity=data.get("current_quantity"),
            minimum_quantity=data.get("minimum_quantity"),
            maximum_quantity=data.get("maximum_quantity"),
            plant_area_location=data.get("plant_area_location"),
            barcode_id=data.get("barcode_id"),
            status=data.get("status", "active"),
            supplier=data.get("supplier"),
            supplier_contact_info=data.get("supplier_contact_info"),
            notes=data.get("notes")
        )

        db.session.add(new_material)
        db.session.commit()

        return jsonify({"message": "Material added successfully", "material": {
            "id": new_material.material_id,
            "title": new_material.title,
            "supplier": new_material.supplier,
            "supplier_contact_info": new_material.supplier_contact_info,
            "notes": new_material.notes
        }}), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
# ➤ Get Material by Barcode
@material_bp.route("/material/barcode/<string:barcode>", methods=["GET"])
def get_material_by_barcode(barcode):
    try:
        material = Material.query.filter_by(barcode_id=barcode).first()

        if not material:
            return jsonify({"message": "Material not found"}), 404

        return jsonify({
            "material_id": material.material_id,
            "title": material.title,
            "description": material.description,
            "unit_of_measure": material.unit_of_measure,
            "current_quantity": material.current_quantity,
            "minimum_quantity": material.minimum_quantity,
            "maximum_quantity": material.maximum_quantity,
            "plant_area_location": material.plant_area_location,
            "barcode_id": material.barcode_id,
            "status": material.status,
            "supplier": material.supplier,
            "supplier_contact_info": material.supplier_contact_info,
            "notes": material.notes
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ➤ Get all Materials
@material_bp.route("/materials", methods=["GET"])
def get_materials():
    materials = Material.query.all()
    return jsonify(materials_schema.dump(materials)), 200

# ➤ Get a specific Material by ID
@material_bp.route("/materials/<int:material_id>", methods=["GET"])
def get_material(material_id):
    material = Material.query.get(material_id)
    if not material:
        return jsonify({"message": "Material not found"}), 404
    return jsonify(material_schema.dump(material)), 200

# ➤ Update a Material
@material_bp.route("/materials/<int:material_id>", methods=["PUT"])
def update_material(material_id):
    material = Material.query.get(material_id)
    if not material:
        return jsonify({"message": "Material not found"}), 404
    
    data = request.get_json()
    material.title = data.get("title", material.title)
    material.description = data.get("description", material.description)
    material.unit_of_measure = data.get("unit_of_measure", material.unit_of_measure)
    material.current_quantity = data.get("current_quantity", material.current_quantity)
    material.minimum_quantity = data.get("minimum_quantity", material.minimum_quantity)
    material.maximum_quantity = data.get("maximum_quantity", material.maximum_quantity)
    material.plant_area_location = data.get("plant_area_location", material.plant_area_location)
    material.barcode_id = data.get("barcode_id", material.barcode_id)
    material.status = data.get("status", material.status)

    db.session.commit()
    return jsonify(material_schema.dump(material)), 200

# ➤ Delete a Material
@material_bp.route("/materials/<int:material_id>", methods=["DELETE"])
def delete_material(material_id):
    # First, delete all transactions related to this material
    MaterialTransaction.query.filter_by(material_id=material_id).delete()

    # Now, delete the material itself
    material = Material.query.get(material_id)
    if not material:
        return jsonify({"error": "Material not found"}), 404
    
    db.session.delete(material)
    db.session.commit()
    
    return jsonify({"message": "Material and associated transactions deleted successfully"}), 200

# ➤ Create a Material Transaction
@material_bp.route("/material-transactions", methods=["POST"])
def create_material_transaction():
    data = request.get_json()
    new_transaction = MaterialTransaction(
        material_id=data["material_id"],
        transaction_type=data["transaction_type"],
        quantity=data["quantity"],
        description=data.get("description")
    )
    db.session.add(new_transaction)
    db.session.commit()
    return jsonify(transaction_schema.dump(new_transaction)), 201

# ➤ Get all Material Transactions
@material_bp.route("/material-transactions", methods=["GET"])
def get_material_transactions():
    transactions = MaterialTransaction.query.all()
    return jsonify(transactions_schema.dump(transactions)), 200

# ➤ Get a specific Material Transaction by ID
@material_bp.route("/material-transactions/<int:transaction_id>", methods=["GET"])
def get_material_transaction(transaction_id):
    transaction = MaterialTransaction.query.get(transaction_id)
    if not transaction:
        return jsonify({"message": "Transaction not found"}), 404
    return jsonify(transaction_schema.dump(transaction)), 200
